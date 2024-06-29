#
# Project Name: Mail Sender
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Lesser General Public License as published
# by the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the
# GNU Lesser General Public License for more details.
#
# You should have received a copy of the GNU Lesser General Public License
# along with this program in the COPYING files.
# If not, see <http://www.gnu.org/licenses/>.
#
# Copyright (C) 2024 Mister Funtikov
#

import os
import re
import sys
import json
import datetime

from PyQt6 import uic
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QObject
from PyQt6.QtWidgets import QApplication, QMainWindow, QMessageBox, QFileDialog

from openpyxl import load_workbook
# from openpyxl import Workbook
# from openpyxl.utils import get_column_letter, column_index_from_string
# from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

from modules.mail_check import mailCheck
from modules.mail_send import mailSend

# -----------------------------------------------


def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)


class OutputLogger(QObject):
    emit_write = pyqtSignal(str, dict)

    class Severity:
        DEBUG = 0
        ERROR = 1

    def __init__(self, io_stream, severity):
        super().__init__()
        self.io_stream = io_stream
        self.severity = severity

    def write(self, text='', params={}):
        # self.io_stream.write(text)
        # self.emit_write.emit(text=text, severity=self.severity)
        self.emit_write.emit(text, params)

    def flush(self):
        # self.io_stream.flush()
        pass


LOGGER = OutputLogger(sys.stdout, OutputLogger.Severity.DEBUG)
LOGGER_STDERR = OutputLogger(sys.stderr, OutputLogger.Severity.ERROR)

sys.stdout = LOGGER
sys.stderr = LOGGER_STDERR

SENDER = mailSend()

# -----------------------------------------------


class SendThread(QThread):

    f_signal = pyqtSignal(str)
    is_on = False

    send_colnum = {
        'ID': None,
        'TO': None,
        'FROM': None,
        'FILENAME': None,
        'TITLE': None
    }

    check = {
        'mail': False,
        'double': False,
        'subfolders': False,
        'attach': False,
        'ssl': False,
    }

    PARAM = {}
    ID_METHOD = 0
    SEND_DIR = None
    SEND_FILE = None
    SEND_BODY = None
    SEND_CONFIG = []

    def __init__(self):
        super(SendThread, self).__init__()
        self.is_on = False

    def preStart(self):

        if self.SEND_CONFIG == None or len(self.SEND_CONFIG) == 0:
            m = 'Не загружен файл парметров рассылки.'
            LOGGER.write(m, params={'status': 'red', 'event': True})
            return False

        if self.ID_METHOD == 1 and self.SEND_FILE != None and self.SEND_FILE != '':
            if os.path.isfile(self.SEND_FILE) == False:
                m = f'Не найден файл: {self.SEND_FILE}'
                LOGGER.write(m, params={'status': 'red', 'event': True})
                return False

        if 'ToEmail' in self.SEND_CONFIG[0]:
            self.send_colnum['TO'] = self.SEND_CONFIG[0].index('ToEmail')
        else:
            m = 'Отсутствует столбец "ToEmail" в файле параметров.'
            LOGGER.write(m, params={'status': 'red', 'event': True})
            return False

        if 'FromEmail' in self.SEND_CONFIG[0]:
            self.send_colnum['FROM'] = self.SEND_CONFIG[0].index('FromEmail')
        else:
            m = 'Отсутствует столбец "FromEmail" в файле параметров.'
            LOGGER.write(m, params={'status': 'red', 'event': True})
            return False

        if 'ID' in self.SEND_CONFIG[0]:
            self.send_colnum['ID'] = self.SEND_CONFIG[0].index('ID')
        else:
            m = 'Отсутствует столбец "ID" в файле параметров.'
            LOGGER.write(m, params={'status': 'red', 'event': True})
            return False

        if 'FileName' in self.SEND_CONFIG[0]:
            self.send_colnum['FILENAME'] = self.SEND_CONFIG[0].index(
                'FileName')

        if 'Title' in self.SEND_CONFIG[0]:
            self.send_colnum['TITLE'] = self.SEND_CONFIG[0].index('Title')

        st = SENDER.connect(self.PARAM['smtp']['host'], self.PARAM['smtp']['port'],
                            self.PARAM['smtp']['user'], self.PARAM['smtp']['password'],
                            self.check['ssl'])
        if st['status'] == False:
            m = 'Не удалось соединиться с сервером. ' + st['msg']
            LOGGER.write(m, params={'status': 'red', 'event': True})
            return False

        if self.SEND_BODY == None or type(self.SEND_BODY) != str:
            self.SEND_BODY = ''

        return True

    def run(self):

        if self.preStart() == False:
            self.sleep(1)
            self.f_signal.emit('false')
            return

        for num, row in enumerate(self.SEND_CONFIG[1:], 1):
            if self.is_on == False:
                break
            self.ssf(num, row)

        self.sleep(1)
        st = SENDER.close()
        self.f_signal.emit('true')
        pass

    def ssf(self, num, row):

        # print(self.send_colnum)
        # {'TO': 2, 'FROM': 3, 'FILENAME': 0, 'TITLE': 1, 'ID': 0}

        mfrom = row[self.send_colnum['FROM']]
        mto = row[self.send_colnum['TO']]

        if self.send_colnum['ID'] != None:
            if row[self.send_colnum['ID']] == '':
                m = f'Пустое поле ID'
                LOGGER.write(
                    m, params={'status': 'red', 'id': f'Строка: {num}'})
                return
            else:
                id = row[self.send_colnum['ID']]
        else:
            id = f'Строка: {num}'

        if mfrom == None or mfrom == '':
            m = f'Отсутствует значение в поле FromEmail'
            LOGGER.write(m, params={'status': 'red', 'id': id})
            return
        if mto == None or mto == '':
            m = f'Отсутствует значение в поле ToEmail'
            LOGGER.write(m, params={'status': 'red', 'id': id})
            return

        if self.check['mail']:
            m = mailCheck(mfrom)
            if m['status'] == False:
                m = f'Неверная почта FromEmail'
                LOGGER.write(m, params={'status': 'red', 'id': id})
                return
            m = mailCheck(
                mto, {'split_symbol': self.PARAM['mail_separator'], 'multiple_mail': True})
            if m['status'] == False:
                m = f'Неверная почта ToEmail'
                LOGGER.write(m, params={'status': 'red', 'id': id})
                return

        mto = re.split(self.PARAM['mail_separator'], mto)

        if self.check['double'] == True:
            mto.append(mfrom)

        if not 'TITLE' in self.send_colnum or self.send_colnum['TITLE'] == '':
            mtitle = 'Электронное письмо'
        else:
            mtitle = row[self.send_colnum['TITLE']]

        mfiles = []

        def viewFiles(root: str, files: list, filename: str) -> list:
            res = []
            if filename == '' or type(filename) != str:
                return []
            for file in files:
                if file.startswith(filename + '_') or file.startswith(filename + ' '):
                    res.append(str(root)+'/'+str(file))
                # if file.endwith('.pdf'):
                #     pass
            return res

        if self.ID_METHOD == 0 and self.SEND_DIR != None:

            # поиск файлов по папке
            if os.path.exists(self.SEND_DIR) == False:
                m = f'Не найдена папка: {self.SEND_DIR}'
                LOGGER.write(m, params={'status': 'red', 'id': id})
                return

            if self.check['subfolders'] == False:
                # r=root, d=directories, f = files
                for r, d, f in os.walk(self.SEND_DIR):
                    mfiles = viewFiles(r, f, id)
                    break
            else:
                # r=root, d=directories, f = files
                for r, d, f in os.walk(self.SEND_DIR):
                    mfiles += viewFiles(r, f, id)

        elif self.ID_METHOD == 1 and self.SEND_FILE != None and self.SEND_FILE != '':
            if os.path.isfile(self.SEND_FILE):
                mfiles = [self.SEND_FILE]
            else:
                m = f'Не найден файл: {self.SEND_FILE}'
                LOGGER.write(m, params={'status': 'red', 'id': id})
                return

        if self.check['attach'] == True and len(mfiles) == 0:
            m = f'Нет вложений для отправки.'
            LOGGER.write(m, params={'status': 'red', 'id': id})
            return

        m = SENDER.send(mfrom=mfrom, mto=mto,
                        title=mtitle, msg=self.SEND_BODY,
                        files=mfiles)
        if m['status'] == False:
            m = f'Не удалось отправить письмо. {m["msg"]}'
            LOGGER.write(m, params={'status': 'red', 'id': id})
            return
        else:
            m = f'Письмо отправлено.'
            LOGGER.write(m, params={'status': 'green', 'id': id})
        return


class MainWindow(QMainWindow):

    PARAM = None
    send_thread = None
    log_filename = None

    def __init__(self):
        super().__init__()
        uic.loadUi(resource_path('design/main.ui'), self)
        self.connObject()
        self.checkConfig()
        # self.form_thread.f_signal.connect(self.set_label_func)
        pass

    def checkConfig(self):

        filename = str(os.getcwd()) + '\config.json'

        if os.path.isfile(filename) == False:
            m = 'Отсутствует файл config.json'
            LOGGER.write(m, params={'status': 'red', 'event': True})
            sys.exit()

        try:
            with open(filename, 'r', encoding='utf-8') as f:
                self.PARAM = json.load(f)
        except Exception as Err:
            m = 'Не удалось прочитать config.json'
            LOGGER.write(m, params={'status': 'red', 'event': True})
            sys.exit()

        if not 'smtp' in self.PARAM:
            m = 'Отсутствует smtp настройка в config.json'
            LOGGER.write(m, params={'status': 'red', 'event': True})
            sys.exit()
        elif not 'host' in self.PARAM['smtp'] or self.PARAM['smtp']['host'] == '':
            m = 'Отсутствует smtp настройка host в config.json'
            LOGGER.write(m, params={'status': 'red', 'event': True})
            sys.exit()
        if not 'port' in self.PARAM['smtp'] or self.PARAM['smtp']['port'] == '':
            self.PARAM['smtp']['port'] = '25'
            m = 'Отсутствует настройка smtp-port в config.json, значение установлено по умолчанию: 25'
            LOGGER.write(m, params={'status': 'yellow'})
        if not 'user' in self.PARAM['smtp'] or self.PARAM['smtp']['user'] == '':
            self.PARAM['smtp']['user'] = 'user'
            m = 'Отсутствует настройка smtp-user в config.json, значение установлено по умолчанию: user'
            LOGGER.write(m, params={'status': 'yellow'})
        if not 'password' in self.PARAM['smtp']:
            self.PARAM['smtp']['password'] = ''

        if not 'default_folder' in self.PARAM:
            self.PARAM['default_folder'] = ''
        if not 'log_filename' in self.PARAM:
            self.PARAM['log_filename'] = f'sendlog_%datetime%'

        if 'use_ssl' in self.PARAM and self.PARAM['use_ssl'] == 'yes':
            self.check_ssl.setChecked(True)
        if 'save_log' in self.PARAM and self.PARAM['save_log'] == 'yes':
            self.check_savelog.setChecked(True)
        if 'use_subfolders' in self.PARAM and self.PARAM['use_subfolders'] == 'yes':
            self.check_subfolders.setChecked(True)
        if 'duplicate_to_sender' in self.PARAM and self.PARAM['duplicate_to_sender'] == 'yes':
            self.check_duble.setChecked(True)
        if 'check_mail' in self.PARAM and self.PARAM['check_mail'] == 'yes':
            self.check_mail.setChecked(True)
        if not 'mail_separator' in self.PARAM or self.PARAM['mail_separator'] == '':
            self.PARAM['mail_separator'] = '; |;|, |,'
        pass

    def connObject(self):
        LOGGER.emit_write.connect(self.setLog)
        LOGGER_STDERR.emit_write.connect(self.setLog)
        self.send_thread = SendThread()
        self.send_thread.f_signal.connect(self.stopSend)
        self.btn_dir_select.clicked.connect(self.dirSelect)
        self.btn_body_select.clicked.connect(self.bodySelect)
        self.btn_param_select.clicked.connect(self.paramSelect)
        self.btn_start.clicked.connect(self.startSend)
        self.btn_stop.clicked.connect(self.stopSend)
        self.combo_method.currentIndexChanged.connect(self.updateMethod)
        pass

    def dirSelect(self):
        if self.send_thread.ID_METHOD == 0:
            dir = str(QFileDialog.getExistingDirectory(
                self, "Укажите папку", directory=self.PARAM['default_folder']))
        else:
            dir = str(QFileDialog.getOpenFileName(
                self, "Укажите файл", filter='', directory=self.PARAM['default_folder'])[0])

        if dir == '':
            return False
        else:
            if self.send_thread.ID_METHOD == 0:
                self.send_thread.SEND_DIR = dir
                m = 'Установлена папка для рассылки'
                LOGGER.write(m, params={'status': 'green'})
            else:
                self.send_thread.SEND_FILE = dir
                m = 'Установлен файл для рассылки'
                LOGGER.write(m, params={'status': 'green'})

        self.edit_dir_path.setText(dir)
        pass

    def bodySelect(self):
        dir = str(QFileDialog.getOpenFileName(
            self, "Укажите файл", filter='(*.htm *.html)', directory=self.PARAM['default_folder'])[0])
        if dir == '':
            return False
        try:
            with open(dir, 'r', encoding='utf-8') as f:
                self.send_thread.SEND_BODY = f.read()
        except Exception as Err:
            m = 'Не удалось прочитать файл'
            LOGGER.write(m, params={'status': 'red', 'event': True})
            return False
        self.edit_body_path.setText(dir)
        m = 'Установлен файл для текста сообщения'
        LOGGER.write(m, params={'status': 'green'})
        return True

    def paramSelect(self):
        dir = str(QFileDialog.getOpenFileName(
            self, "Укажите файл с параметрами", filter='(*.xlsx)', directory=self.PARAM['default_folder'])[0])
        if dir == '':
            return False
        try:
            wb = load_workbook(dir)
            ws = wb[wb.sheetnames[0]]
        except Exception as Err:
            m = 'Файл с параметрами не загружен'
            LOGGER.write(m, params={'status': 'red', 'event': True})
            return False

        self.edit_param.setText(dir)
        self.send_thread.SEND_CONFIG = []

        for row in tuple(ws.iter_rows(values_only=True)):
            self.send_thread.SEND_CONFIG.append(row)

        m = 'Файл с параметрами загружен'
        LOGGER.write(m, params={'status': 'green'})
        return True

    def updateMethod(self):
        # 0 - many files - many address
        # 1 - one files - many address
        mid = self.combo_method.currentIndex()
        self.edit_dir_path.setText('')
        self.send_thread.SEND_DIR = None
        self.send_thread.SEND_FILE = None
        self.send_thread.ID_METHOD = mid
        if mid == 0:
            self.label_methoddir.setText('Папка:')
            self.check_subfolders.setEnabled(True)
        else:
            self.label_methoddir.setText('Файл:')
            self.check_subfolders.setEnabled(False)
            self.check_subfolders.setChecked(False)
        pass

    def setLog(self, text='', params={}):

        color = '#000000'
        tp = 'инфо'
        id = '-'

        if type(params) == dict:
            if 'status' in params:
                if params['status'] == 'red':
                    color = '#D51B21'
                    tp = 'ошибка'
                elif params['status'] == 'yellow':
                    color = '#FFA733'
                    tp = 'внимание'
                elif params['status'] == 'green':
                    color = '#008000'
                    tp = 'успех'
                else:
                    color = '#000000'
                    tp = 'инфо'
            if 'id' in params and params['id'] != '':
                id = params['id']

        tx = f'{id}\t{tp}\t{datetime.datetime.now().strftime("%d.%m.%Y")}\t{datetime.datetime.now().strftime("%H:%M:%S")}\t{text}'
        ltx = f'<span style="color:{color}">{tx}</span>'

        self.text_log.append(ltx)

        if 'linebreak' in params and params['linebreak'] != False:
            self.text_log.append('<br/>')

        if 'event' in params and params['event'] == True:
            self.errorEvent(text)

        if self.check_savelog.isChecked():
            if self.log_filename == None or self.log_filename == '':
                self.log_filename = 'sendlog_' + \
                    str(datetime.datetime.now().strftime(
                        '%d-%m-%Y_%H-%M-%S')) + '.txt'
            try:
                f = open(self.log_filename, 'a', encoding='UTF-8')
                f.write(tx+'\n')
                f.close()
            except Exception as Err:
                pass

        pass

    def errorEvent(self, errtext):
        reply = QMessageBox.critical(self, 'Ошибка', errtext,
                                     QMessageBox.StandardButton.Close,
                                     QMessageBox.StandardButton.Close)
        if reply:
            pass

    def closeEvent(self, event):
        reply = QMessageBox.question(self, 'Закрытие программы', 'Действительно закрыть программу?',
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                                     QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            event.accept()
        else:
            event.ignore()
        pass

    def lockForm(self):
        self.btn_param_select.setEnabled(False)
        self.btn_body_select.setEnabled(False)
        self.combo_method.setEnabled(False)
        self.btn_dir_select.setEnabled(False)
        self.btn_start.setEnabled(False)
        self.paramBox.setEnabled(False)
        self.btn_stop.setEnabled(True)
        pass

    def unlockForm(self):
        self.btn_param_select.setEnabled(True)
        self.btn_body_select.setEnabled(True)
        self.combo_method.setEnabled(True)
        self.btn_dir_select.setEnabled(True)
        self.btn_start.setEnabled(True)
        self.paramBox.setEnabled(True)
        self.btn_stop.setEnabled(False)
        pass

    def stopSend(self, status='true'):
        self.send_thread.is_on = False
        self.unlockForm()
        pass

    def startSend(self):

        self.lockForm()

        if self.check_mail.isChecked():
            self.send_thread.check['mail'] = True
        else:
            self.send_thread.check['mail'] = False

        if self.check_duble.isChecked():
            self.send_thread.check['double'] = True
        else:
            self.send_thread.check['double'] = False

        if self.check_subfolders.isChecked():
            self.send_thread.check['subfolders'] = True
        else:
            self.send_thread.check['subfolders'] = False

        if self.check_attach.isChecked():
            self.send_thread.check['attach'] = True
        else:
            self.send_thread.check['attach'] = False

        if self.check_ssl.isChecked():
            self.send_thread.check['ssl'] = True
        else:
            self.send_thread.check['ssl'] = False

        self.send_thread.PARAM = self.PARAM

        self.send_thread.is_on = True
        self.send_thread.start()
        pass


if __name__ == '__main__':

    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())
